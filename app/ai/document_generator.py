"""
Document generator for creating missing checklist documents
"""

from pathlib import Path
from typing import Dict, Any, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


class DocumentGenerator:
    """Generates accounting documents based on templates and user data"""
    
    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_bank_reconciliation(
        self, 
        data: Dict[str, Any], 
        filename: Optional[str] = None
    ) -> Path:
        """Generate bank reconciliation template"""
        
        if filename is None:
            filename = f"bank_reconciliation_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = self.output_dir / filename
        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Reconciliation"
        
        # Header
        ws['A1'] = "BANK RECONCILIATION"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Period: {data.get('period', 'Month End')}"
        ws['A3'] = f"Bank Account: {data.get('bank_account', '')}"
        
        # Reconciliation structure
        row = 5
        ws[f'A{row}'] = "Balance per Bank Statement"
        ws[f'C{row}'] = data.get('bank_balance', 0)
        
        row += 2
        ws[f'A{row}'] = "Add: Deposits in Transit"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        deposits = data.get('deposits_in_transit', [])
        for deposit in deposits[:5]:  # Max 5 sample rows
            ws[f'A{row}'] = deposit.get('description', '')
            ws[f'C{row}'] = deposit.get('amount', 0)
            row += 1
        
        # Add a few empty rows for user to fill
        for _ in range(3):
            ws[f'A{row}'] = ''
            ws[f'C{row}'] = 0
            row += 1
        
        ws[f'A{row}'] = "Total Deposits in Transit"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f"=SUM(C{row-6}:C{row-1})"
        
        row += 2
        ws[f'A{row}'] = "Less: Outstanding Checks"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        checks = data.get('outstanding_checks', [])
        for check in checks[:5]:
            ws[f'A{row}'] = f"Check # {check.get('check_no', '')}"
            ws[f'C{row}'] = check.get('amount', 0)
            row += 1
        
        for _ in range(3):
            ws[f'A{row}'] = ''
            ws[f'C{row}'] = 0
            row += 1
        
        ws[f'A{row}'] = "Total Outstanding Checks"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f"=SUM(C{row-6}:C{row-1})"
        
        row += 2
        ws[f'A{row}'] = "Balance per Books"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'C{row}'] = f"=C5+C{row-10}-C{row-1}"
        ws[f'C{row}'].font = Font(bold=True)
        
        # Formatting
        self._apply_basic_formatting(ws)
        
        wb.save(filepath)
        return filepath
    
    def generate_ar_aging(
        self, 
        data: Dict[str, Any], 
        filename: Optional[str] = None
    ) -> Path:
        """Generate AR aging report template"""
        
        if filename is None:
            filename = f"ar_aging_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = self.output_dir / filename
        
        # Create DataFrame
        columns = [
            'Customer Name',
            'Invoice Number',
            'Invoice Date',
            'Total Amount',
            'Current (0-30)',
            '31-60 Days',
            '61-90 Days',
            '90+ Days'
        ]
        
        # Add sample data if provided
        customers = data.get('customers', [])
        rows = []
        
        for customer in customers:
            rows.append({
                'Customer Name': customer.get('name', ''),
                'Invoice Number': customer.get('invoice_no', ''),
                'Invoice Date': customer.get('invoice_date', ''),
                'Total Amount': customer.get('amount', 0),
                'Current (0-30)': customer.get('current', 0),
                '31-60 Days': customer.get('30_days', 0),
                '61-90 Days': customer.get('60_days', 0),
                '90+ Days': customer.get('90_days', 0)
            })
        
        # Add empty rows for user to fill
        for i in range(10):
            rows.append({col: '' if 'Name' in col or 'Number' in col or 'Date' in col else 0 
                        for col in columns})
        
        df = pd.DataFrame(rows, columns=columns)
        
        # Save to Excel with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='AR Aging', index=False)
            
            worksheet = writer.sheets['AR Aging']
            self._apply_basic_formatting(worksheet)
            
            # Add total row
            last_row = len(df) + 2
            worksheet[f'A{last_row}'] = 'TOTAL'
            worksheet[f'A{last_row}'].font = Font(bold=True)
            for col_idx in range(4, 9):  # Columns E to H
                col_letter = chr(64 + col_idx)
                worksheet[f'{col_letter}{last_row}'] = f'=SUM({col_letter}2:{col_letter}{last_row-1})'
                worksheet[f'{col_letter}{last_row}'].font = Font(bold=True)
        
        return filepath
    
    def generate_ap_aging(
        self, 
        data: Dict[str, Any], 
        filename: Optional[str] = None
    ) -> Path:
        """Generate AP aging report template"""
        
        if filename is None:
            filename = f"ap_aging_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = self.output_dir / filename
        
        columns = [
            'Vendor Name',
            'Invoice Number',
            'Invoice Date',
            'Total Amount',
            'Current (0-30)',
            '31-60 Days',
            '61-90 Days',
            '90+ Days'
        ]
        
        vendors = data.get('vendors', [])
        rows = []
        
        for vendor in vendors:
            rows.append({
                'Vendor Name': vendor.get('name', ''),
                'Invoice Number': vendor.get('invoice_no', ''),
                'Invoice Date': vendor.get('invoice_date', ''),
                'Total Amount': vendor.get('amount', 0),
                'Current (0-30)': vendor.get('current', 0),
                '31-60 Days': vendor.get('30_days', 0),
                '61-90 Days': vendor.get('60_days', 0),
                '90+ Days': vendor.get('90_days', 0)
            })
        
        for i in range(10):
            rows.append({col: '' if 'Name' in col or 'Number' in col or 'Date' in col else 0 
                        for col in columns})
        
        df = pd.DataFrame(rows, columns=columns)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='AP Aging', index=False)
            worksheet = writer.sheets['AP Aging']
            self._apply_basic_formatting(worksheet)
            
            last_row = len(df) + 2
            worksheet[f'A{last_row}'] = 'TOTAL'
            worksheet[f'A{last_row}'].font = Font(bold=True)
            for col_idx in range(4, 9):
                col_letter = chr(64 + col_idx)
                worksheet[f'{col_letter}{last_row}'] = f'=SUM({col_letter}2:{col_letter}{last_row-1})'
                worksheet[f'{col_letter}{last_row}'].font = Font(bold=True)
        
        return filepath
    
    def generate_accrual_journal(
        self, 
        data: Dict[str, Any], 
        filename: Optional[str] = None
    ) -> Path:
        """Generate accrual journal entries template"""
        
        if filename is None:
            filename = f"accrual_journal_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = self.output_dir / filename
        
        columns = [
            'JE Number',
            'Date',
            'Account Code',
            'Account Name',
            'Description',
            'Debit',
            'Credit'
        ]
        
        entries = data.get('entries', [])
        rows = []
        
        for entry in entries:
            rows.append({
                'JE Number': entry.get('je_number', ''),
                'Date': entry.get('date', ''),
                'Account Code': entry.get('account_code', ''),
                'Account Name': entry.get('account_name', ''),
                'Description': entry.get('description', ''),
                'Debit': entry.get('debit', 0),
                'Credit': entry.get('credit', 0)
            })
        
        for i in range(15):
            rows.append({col: '' if 'Number' in col or 'Date' in col or 'Code' in col or 'Name' in col or 'Description' in col else 0 
                        for col in columns})
        
        df = pd.DataFrame(rows, columns=columns)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Journal Entries', index=False)
            worksheet = writer.sheets['Journal Entries']
            self._apply_basic_formatting(worksheet)
            
            last_row = len(df) + 2
            worksheet[f'E{last_row}'] = 'TOTAL'
            worksheet[f'E{last_row}'].font = Font(bold=True)
            worksheet[f'F{last_row}'] = f'=SUM(F2:F{last_row-1})'
            worksheet[f'G{last_row}'] = f'=SUM(G2:G{last_row-1})'
            worksheet[f'F{last_row}'].font = Font(bold=True)
            worksheet[f'G{last_row}'].font = Font(bold=True)
            
            # Add validation note
            worksheet[f'A{last_row+2}'] = 'Note: Debits must equal Credits'
            worksheet[f'A{last_row+3}'] = 'Difference:'
            worksheet[f'B{last_row+3}'] = f'=F{last_row}-G{last_row}'
        
        return filepath
    
    def generate_expense_analysis(
        self,
        data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> Path:
        """Generate expense analysis template"""
        
        if filename is None:
            filename = f"expense_analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = self.output_dir / filename
        
        columns = [
            'Date',
            'Expense Category',
            'Sub-Category',
            'Description',
            'Amount',
            'Budget',
            'Variance',
            'Variance %',
            'Department',
            'Notes'
        ]
        
        # Add sample data if provided
        expenses = data.get('expenses', [])
        rows = []
        
        for expense in expenses:
            rows.append({
                'Date': expense.get('date', ''),
                'Expense Category': expense.get('category', ''),
                'Sub-Category': expense.get('sub_category', ''),
                'Description': expense.get('description', ''),
                'Amount': expense.get('amount', 0),
                'Budget': expense.get('budget', 0),
                'Variance': expense.get('variance', 0),
                'Variance %': expense.get('variance_pct', 0),
                'Department': expense.get('department', ''),
                'Notes': expense.get('notes', '')
            })
        
        # Add empty rows for user to fill
        for i in range(20):
            rows.append({
                'Date': '',
                'Expense Category': '',
                'Sub-Category': '',
                'Description': '',
                'Amount': 0,
                'Budget': 0,
                'Variance': '=F{}-E{}'.format(len(expenses) + i + 2, len(expenses) + i + 2),
                'Variance %': '=IF(F{}=0,0,G{}/F{})'.format(len(expenses) + i + 2, len(expenses) + i + 2, len(expenses) + i + 2),
                'Department': '',
                'Notes': ''
            })
        
        df = pd.DataFrame(rows, columns=columns)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Expense Analysis', index=False)
            worksheet = writer.sheets['Expense Analysis']
            self._apply_basic_formatting(worksheet)
            
            # Add total row
            last_row = len(df) + 2
            worksheet[f'D{last_row}'] = 'TOTAL'
            worksheet[f'D{last_row}'].font = Font(bold=True)
            worksheet[f'E{last_row}'] = f'=SUM(E2:E{last_row-1})'
            worksheet[f'F{last_row}'] = f'=SUM(F2:F{last_row-1})'
            worksheet[f'G{last_row}'] = f'=SUM(G2:G{last_row-1})'
            worksheet[f'E{last_row}'].font = Font(bold=True)
            worksheet[f'F{last_row}'].font = Font(bold=True)
            worksheet[f'G{last_row}'].font = Font(bold=True)
        
        return filepath
    
    def generate_prepayments_schedule(
        self,
        data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> Path:
        """Generate prepayments schedule template"""
        
        if filename is None:
            filename = f"prepayments_schedule_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = self.output_dir / filename
        
        columns = [
            'Prepayment Type',
            'Vendor/Description',
            'Start Date',
            'End Date',
            'Total Amount',
            'Monthly Amortization',
            'Current Month',
            'Remaining Balance',
            'Account Code'
        ]
        
        df = pd.DataFrame(columns=columns)
        for i in range(15):
            df.loc[i] = [''] * len(columns)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Prepayments', index=False)
            worksheet = writer.sheets['Prepayments']
            self._apply_basic_formatting(worksheet)
        
        return filepath
    
    def generate_revenue_schedule(
        self,
        data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> Path:
        """Generate revenue schedule template"""
        
        if filename is None:
            filename = f"revenue_schedule_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = self.output_dir / filename
        
        columns = [
            'Revenue Stream',
            'Customer/Source',
            'Invoice Number',
            'Invoice Date',
            'Amount',
            'Recognition Period',
            'Current Month Revenue',
            'Deferred Revenue',
            'Account Code'
        ]
        
        df = pd.DataFrame(columns=columns)
        for i in range(20):
            df.loc[i] = [''] * len(columns)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Revenue Schedule', index=False)
            worksheet = writer.sheets['Revenue Schedule']
            self._apply_basic_formatting(worksheet)
        
        return filepath
    
    def generate_fixed_assets_register(
        self,
        data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> Path:
        """Generate fixed assets register template"""
        
        if filename is None:
            filename = f"fixed_assets_register_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = self.output_dir / filename
        
        columns = [
            'Asset ID',
            'Asset Description',
            'Category',
            'Purchase Date',
            'Cost',
            'Accumulated Depreciation',
            'Net Book Value',
            'Depreciation Method',
            'Useful Life (Years)',
            'Current Month Depreciation',
            'Location'
        ]
        
        df = pd.DataFrame(columns=columns)
        for i in range(15):
            df.loc[i] = [''] * len(columns)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Fixed Assets', index=False)
            worksheet = writer.sheets['Fixed Assets']
            self._apply_basic_formatting(worksheet)
        
        return filepath
    
    def generate_intercompany_reconciliation(
        self,
        data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> Path:
        """Generate intercompany reconciliation template"""
        
        if filename is None:
            filename = f"intercompany_recon_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = self.output_dir / filename
        
        columns = [
            'Entity A',
            'Entity B',
            'Transaction Description',
            'Date',
            'Amount (Entity A)',
            'Amount (Entity B)',
            'Difference',
            'Reconciliation Status',
            'Notes'
        ]
        
        df = pd.DataFrame(columns=columns)
        for i in range(15):
            df.loc[i] = [''] * len(columns)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Intercompany Recon', index=False)
            worksheet = writer.sheets['Intercompany Recon']
            self._apply_basic_formatting(worksheet)
        
        return filepath
    
    def generate_generic_template(
        self, 
        checklist_item_id: str,
        columns: List[str],
        filename: Optional[str] = None
    ) -> Path:
        """Generate a generic template for any checklist item"""
        
        if filename is None:
            filename = f"{checklist_item_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = self.output_dir / filename
        
        # If no columns provided, use defaults
        if not columns or columns == ["Column1", "Column2", "Column3"]:
            columns = ["Date", "Description", "Amount", "Category", "Notes"]
        
        # Create empty DataFrame with specified columns
        df = pd.DataFrame(columns=columns)
        
        # Add 20 empty rows
        for i in range(20):
            df.loc[i] = ['' if j < len(columns)-1 else 0 for j in range(len(columns))]
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)
            worksheet = writer.sheets['Data']
            self._apply_basic_formatting(worksheet)
        
        return filepath
    
    def _apply_basic_formatting(self, worksheet):
        """Apply basic formatting to Excel worksheet"""
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
